# app/api/products.py
from __future__ import annotations

from typing import Optional, List, Dict, Any, Tuple
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Body, Depends, Request
from fastapi.responses import Response, FileResponse
from pydantic import BaseModel
from contextlib import contextmanager
from dataclasses import asdict, is_dataclass, dataclass
import io
import os
import shutil
import sqlite3
import datetime as _dt
from app.kaspi_client import KaspiClient
from app.deps.kaspi_client import get_kaspi_client
# ──────────────────────────────────────────────────────────────────────────────
# optional deps
# ──────────────────────────────────────────────────────────────────────────────
try:
    import openpyxl
    _OPENPYXL_OK = True
except Exception:
    _OPENPYXL_OK = False

try:
    import requests  # для XML-фида Kaspi
    _REQ_OK = True
except Exception:
    _REQ_OK = False

# ──────────────────────────────────────────────────────────────────────────────
# DB backends (PG via SQLAlchemy / fallback SQLite)
# ──────────────────────────────────────────────────────────────────────────────
try:
    from sqlalchemy import create_engine, text
    _SQLA_OK = True
except Exception:
    _SQLA_OK = False

DATABASE_URL = os.getenv("DATABASE_URL")  # e.g. postgresql+psycopg://...
_USE_PG = bool(DATABASE_URL and _SQLA_OK)

def _resolve_db_path() -> str:
    target = os.getenv("DB_PATH", "/data/kaspi-orders.sqlite3")
    d = os.path.dirname(target)
    try:
        os.makedirs(d, exist_ok=True)
        if os.access(d, os.W_OK):
            return target
    except Exception:
        pass
    return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data.sqlite3"))

DB_PATH = _resolve_db_path()

# migrate old local file if path changed
_OLD_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data.sqlite3"))
if DB_PATH != _OLD_PATH and os.path.exists(_OLD_PATH) and not os.path.exists(DB_PATH):
    try:
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
        shutil.copy2(_OLD_PATH, DB_PATH)
    except Exception:
        pass

if _USE_PG:
    _engine = create_engine(DATABASE_URL, pool_pre_ping=True, future=True)

@contextmanager
def _db():
    if _USE_PG:
        with _engine.begin() as conn:
            yield conn
    else:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
        finally:
            try:
                conn.commit()
            except Exception:
                pass
            conn.close()

def _q(sql: str):
    return text(sql) if _USE_PG else sql

def _rows_to_dicts(rows):
    if _USE_PG:
        return [dict(r._mapping) for r in rows]
    return [dict(r) for r in rows]

def _commit(c):
    if not _USE_PG:
        try:
            c.commit()
        except Exception:
            pass

# ──────────────────────────────────────────────────────────────────────────────
# Auth (X-API-Key)
# ──────────────────────────────────────────────────────────────────────────────
def _require_api_key(req: Request) -> bool:
    api_key = os.getenv("API_KEY") or os.getenv("KASPI_API_KEY") or os.getenv("PRODUCTS_API_KEY")
    if not api_key:
        return True
    sent = req.headers.get("X-API-Key") or req.query_params.get("api_key")
    if sent != api_key:
        raise HTTPException(401, "Invalid API key")
    return True

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────
def _env_bool(name: str, default: bool) -> bool:
    v = os.getenv(name)
    if v is None: return default
    s = str(v).strip().lower()
    return s in ("1","true","yes","on","+","y")

def _env_float(name: str, default: float) -> float:
    try:
        return float(os.getenv(name, default))
    except Exception:
        return float(default)

def _gen_batch_code() -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    import random as _r
    return "".join(_r.choice(alphabet) for _ in range(6))

def _table_exists(c, name: str) -> bool:
    if _USE_PG:
        r = c.execute(_q("SELECT 1 FROM information_schema.tables WHERE table_name=:t"), {"t": name}).first()
        return bool(r)
    else:
        r = c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (name,)).fetchone()
        return bool(r)

def _has_column(c, table: str, col: str) -> bool:
    if _USE_PG:
        r = c.execute(_q("""SELECT 1 FROM information_schema.columns
                             WHERE table_name=:t AND column_name=:c"""), {"t": table, "c": col}).first()
        return bool(r)
    else:
        rows = c.execute(f"PRAGMA table_info({table})").fetchall()
        return col in {row["name"] for row in rows}

def _maybe_float(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s.replace(" ", "").replace(",", "."))
    except Exception:
        try:
            return float(v)
        except Exception:
            return None

def _maybe_int(v) -> Optional[int]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s.replace(" ", "").replace(",", ".")))
    except Exception:
        try:
            return int(v)
        except Exception:
            return None

def _norm_sku(s: str) -> str:
    if s is None:
        return ""
    return "".join(ch for ch in str(s).strip() if not ch.isspace())

def _sku_of(row: dict) -> str:
    raw = (
        row.get("sku") or row.get("code") or row.get("vendorCode")
        or row.get("barcode") or row.get("id") or ""
    )
    return _norm_sku(raw)

# ──────────────────────────────────────────────────────────────────────────────
# Schema & migrations
# ──────────────────────────────────────────────────────────────────────────────
def _ensure_schema():
    if _USE_PG:
        with _db() as c:
            c.execute(_q("""
            CREATE TABLE IF NOT EXISTS products(
                sku TEXT PRIMARY KEY,
                name TEXT,
                brand TEXT,
                category TEXT,
                price DOUBLE PRECISION,
                quantity INTEGER,
                active INTEGER DEFAULT 1,
                barcode TEXT,
                updated_at TIMESTAMP DEFAULT NOW()
            );"""))
            c.execute(_q("""
            CREATE TABLE IF NOT EXISTS batches(
                id SERIAL PRIMARY KEY,
                sku TEXT NOT NULL REFERENCES products(sku) ON DELETE CASCADE,
                date DATE NOT NULL,
                qty INTEGER NOT NULL,
                unit_cost DOUBLE PRECISION NOT NULL,
                note TEXT,
                commission_pct DOUBLE PRECISION,
                batch_code TEXT UNIQUE,
                qty_sold INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT NOW()
            );"""))
            c.execute(_q("""
            CREATE TABLE IF NOT EXISTS categories(
                name TEXT PRIMARY KEY,
                base_percent DOUBLE PRECISION DEFAULT 0.0,
                extra_percent DOUBLE PRECISION DEFAULT 3.0,
                tax_percent DOUBLE PRECISION DEFAULT 0.0
            );"""))
            c.execute(_q("CREATE INDEX IF NOT EXISTS idx_batches_sku ON batches(sku)"))
            c.execute(_q("CREATE INDEX IF NOT EXISTS idx_batches_date ON batches(date)"))

            if not _has_column(c, "batches", "qty_sold"):
                c.execute(_q("ALTER TABLE batches ADD COLUMN qty_sold INTEGER DEFAULT 0"))

            # cleanup dupes (safety)
            dup = c.execute(_q("""SELECT sku, COUNT(*) AS c
                                  FROM products GROUP BY sku HAVING COUNT(*)>1""")).all()
            if dup:
                c.execute(_q("""
                    DELETE FROM products p USING (
                        SELECT ctid, sku, row_number() OVER (PARTITION BY sku ORDER BY updated_at NULLS LAST) AS rn
                        FROM products
                    ) t
                    WHERE p.ctid=t.ctid AND t.rn>1
                """))
    else:
        with _db() as c:
            c.executescript("""
            PRAGMA journal_mode=WAL;

            CREATE TABLE IF NOT EXISTS products(
                sku TEXT PRIMARY KEY,
                name TEXT,
                brand TEXT,
                category TEXT,
                price REAL,
                quantity INTEGER,
                active INTEGER DEFAULT 1,
                barcode TEXT,
                updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS batches(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sku TEXT NOT NULL,
                date TEXT NOT NULL,
                qty INTEGER NOT NULL,
                unit_cost REAL NOT NULL,
                note TEXT,
                commission_pct REAL,
                batch_code TEXT UNIQUE,
                qty_sold INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS categories(
                name TEXT PRIMARY KEY,
                base_percent REAL DEFAULT 0.0,
                extra_percent REAL DEFAULT 3.0,
                tax_percent REAL DEFAULT 0.0
            );

            CREATE INDEX IF NOT EXISTS idx_batches_sku ON batches(sku);
            CREATE INDEX IF NOT EXISTS idx_batches_date ON batches(date);
            """)
            cols = {r["name"] for r in c.execute("PRAGMA table_info(batches)")}

            if "commission_pct" not in cols:
                c.execute("ALTER TABLE batches ADD COLUMN commission_pct REAL")
            if "qty_sold" not in cols:
                c.execute("ALTER TABLE batches ADD COLUMN qty_sold INTEGER DEFAULT 0")
            if "batch_code" not in cols:
                c.execute("ALTER TABLE batches ADD COLUMN batch_code TEXT")

            # generate missing codes
            for r in c.execute("SELECT id FROM batches WHERE batch_code IS NULL OR batch_code=''").fetchall():
                c.execute("UPDATE batches SET batch_code=? WHERE id=?", (_gen_batch_code(), r["id"]))

            # ensure sku uniqueness
            c.executescript("""
                DELETE FROM products
                 WHERE rowid NOT IN (SELECT MIN(rowid) FROM products GROUP BY sku);
            """)
            _commit(c)

def _seed_categories_if_empty():
    _ensure_schema()
    with _db() as c:
        if _USE_PG:
            cnt = c.execute(_q("SELECT COUNT(*) AS c FROM categories")).scalar_one()
        else:
            cnt = c.execute("SELECT COUNT(*) AS c FROM categories").fetchone()["c"]
        if int(cnt) == 0:
            defaults = [
                ("Витамины/БАДы", 10.0, 3.0, 0.0),
                ("Сад/освещение", 10.0, 3.0, 0.0),
                ("Товары для дома", 10.0, 3.0, 0.0),
                ("Прочее", 10.0, 3.0, 0.0),
            ]
            if _USE_PG:
                for n,b,e,t in defaults:
                    c.execute(_q("""
                        INSERT INTO categories(name,base_percent,extra_percent,tax_percent)
                        VALUES(:n,:b,:e,:t)
                        ON CONFLICT (name) DO NOTHING
                    """), {"n": n, "b": b, "e": e, "t": t})
            else:
                c.executemany(
                    "INSERT OR IGNORE INTO categories(name,base_percent,extra_percent,tax_percent) VALUES(?,?,?,?)",
                    defaults
                )
                _commit(c)

# ──────────────────────────────────────────────────────────────────────────────
# UPSERT & SYNC
# ──────────────────────────────────────────────────────────────────────────────
def _upsert_products(items: List[Dict[str, Any]], *, price_only: bool = True) -> Tuple[int, int]:
    _ensure_schema()
    inserted = updated = 0
    po = 1 if price_only else 0

    with _db() as c:
        for it in items:
            sku = _sku_of(it)
            if not sku:
                continue

            params = {
                "sku": sku,
                "name": (it.get("name") or None),
                "brand": (it.get("brand") or None),
                "category": (it.get("category") or None),
                "price": _maybe_float(it.get("price")),
                "quantity": _maybe_int(it.get("qty") or it.get("quantity") or it.get("stock")),
                "active": (
                    1 if str(it.get("active")).lower() in ("1","true","yes","on","published","active") else
                    0 if str(it.get("active")).lower() in ("0","false","no","off") else
                    None
                ),
                "barcode": (it.get("barcode") or None),
                "price_only": po,
            }

            if _USE_PG:
                existed = c.execute(_q("SELECT 1 FROM products WHERE sku=:sku"), {"sku": sku}).first()
                c.execute(_q("""
                    INSERT INTO products(sku,name,brand,category,price,quantity,active,barcode,updated_at)
                    VALUES(:sku,:name,:brand,:category,:price,:quantity,:active,:barcode,NOW())
                    ON CONFLICT (sku) DO UPDATE SET
                        name     = COALESCE(NULLIF(EXCLUDED.name,''), products.name),
                        brand    = COALESCE(NULLIF(EXCLUDED.brand,''), products.brand),
                        category = COALESCE(NULLIF(EXCLUDED.category,''), products.category),
                        price    = COALESCE(EXCLUDED.price, products.price),
                        quantity = CASE WHEN :price_only = 1 THEN products.quantity
                                        ELSE COALESCE(EXCLUDED.quantity, products.quantity) END,
                        active   = CASE WHEN :price_only = 1 THEN products.active
                                        ELSE COALESCE(EXCLUDED.active, products.active) END,
                        barcode  = COALESCE(NULLIF(EXCLUDED.barcode,''), products.barcode),
                        updated_at = NOW()
                """), params)
            else:
                existed = c.execute("SELECT 1 FROM products WHERE sku=?", (sku,)).fetchone()
                c.execute("""
                    INSERT INTO products(sku,name,brand,category,price,quantity,active,barcode,updated_at)
                    VALUES(?,?,?,?,?,?,?, ?, datetime('now'))
                    ON CONFLICT(sku) DO UPDATE SET
                        name     = CASE WHEN excluded.name     IS NOT NULL AND excluded.name     <> '' THEN excluded.name     ELSE name END,
                        brand    = CASE WHEN excluded.brand    IS NOT NULL AND excluded.brand    <> '' THEN excluded.brand    ELSE brand END,
                        category = CASE WHEN excluded.category IS NOT NULL AND excluded.category <> '' THEN excluded.category ELSE category END,
                        price    = COALESCE(excluded.price,    price),
                        quantity = CASE WHEN ?=1 THEN quantity ELSE COALESCE(excluded.quantity, quantity) END,
                        active   = CASE WHEN ?=1 THEN active   ELSE COALESCE(excluded.active,   active)   END,
                        barcode  = CASE WHEN excluded.barcode  IS NOT NULL AND excluded.barcode  <> '' THEN excluded.barcode  ELSE barcode END,
                        updated_at = datetime('now')
                """, (params["sku"], params["name"], params["brand"], params["category"],
                      params["price"], params["quantity"], params["active"], params["barcode"],
                      po, po))
                _commit(c)

            if existed: updated += 1
            else: inserted += 1

    return inserted, updated

def bulk_upsert_products(rows: List[Dict[str, Any]], *, price_only: bool = True) -> Dict[str, int]:
    inserted, updated = _upsert_products(rows, price_only=price_only)
    return {"inserted": inserted, "updated": updated}

def _count_active_in_db() -> int:
    with _db() as c:
        if _USE_PG:
            r = c.execute(_q("SELECT COUNT(*) AS c FROM products WHERE active=1")).first()
            return int(r._mapping["c"] if r else 0)
        else:
            r = c.execute("SELECT COUNT(*) AS c FROM products WHERE active=1").fetchone()
            return int(r["c"] if r else 0)

def _deactivate_missing(keep_skus: List[str]) -> int:
    if not keep_skus:
        return 0
    with _db() as c:
        if _USE_PG:
            placeholders = ", ".join([f":s{i}" for i in range(len(keep_skus))])
            sql = _q(f"UPDATE products SET active=0 WHERE sku NOT IN ({placeholders}) AND active<>0")
            params = {f"s{i}": s for i, s in enumerate(keep_skus)}
            r = c.execute(sql, params)
            return r.rowcount or 0
        else:
            placeholders = ", ".join(["?"] * len(keep_skus))
            r = c.execute(f"UPDATE products SET active=0 WHERE sku NOT IN ({placeholders}) AND active<>0", keep_skus)
            n = r.rowcount or 0
            _commit(c)
            return n

def _delete_missing(keep_skus: List[str]) -> int:
    if not keep_skus:
        return 0
    with _db() as c:
        if _USE_PG:
            placeholders = ", ".join([f":s{i}" for i in range(len(keep_skus))])
            sql = _q(f"DELETE FROM products WHERE sku NOT IN ({placeholders})")
            params = {f"s{i}": s for i, s in enumerate(keep_skus)}
            r = c.execute(sql, params)
            return r.rowcount or 0
        else:
            placeholders = ", ".join(["?"] * len(keep_skus))
            r = c.execute(f"DELETE FROM products WHERE sku NOT IN ({placeholders})", keep_skus)
            n = r.rowcount or 0
            _commit(c)
            return n

# ──────────────────────────────────────────────────────────────────────────────
# Parsers (Kaspi XML / Excel)
# ──────────────────────────────────────────────────────────────────────────────
def _parse_xml_smart(raw: bytes, *, city_id: str) -> List[Dict[str, Any]]:
    from xml.etree import ElementTree as ET
    try:
        root = ET.fromstring(raw)
    except ET.ParseError as e:
        raise HTTPException(400, f"Некорректный XML: {e}")

    def strip(tag: str) -> str:
        return tag.split("}", 1)[-1] if "}" in tag else tag

    def first_text(parent: ET.Element, *names: str) -> str:
        for el in parent.iter():
            if strip(el.tag) in names:
                t = (el.text or "").strip()
                if t:
                    return t
        return ""

    rows: Dict[str, Dict[str, Any]] = {}
    for off in (el for el in root.iter() if strip(el.tag) == "offer"):
        code = _norm_sku(off.get("sku") or off.get("shop-sku") or off.get("code") or off.get("id") or "")
        if not code:
            continue
        name = first_text(off, "model", "name", "title") or code
        brand = first_text(off, "brand") or None

        # price — prefer city_id
        price = None
        for el in off.iter():
            if strip(el.tag) == "cityprice" and (el.get("cityId") or "") == city_id:
                price = _maybe_float(el.text)
                break
        if price is None:
            for el in off.iter():
                if strip(el.tag) == "cityprice":
                    price = _maybe_float(el.text); break
        if price is None:
            p = first_text(off, "price")
            price = _maybe_float(p)

        qty, active = None, None
        for el in off.iter():
            if strip(el.tag) == "availability":
                sc = el.get("stockCount")
                qty = _maybe_int(sc) if sc is not None else None
                av = (el.get("available") or "").strip().lower()
                active = True if av in ("yes","true","1") else False if av in ("no","false","0") else None
                break

        rows[code] = {
            "sku": code, "code": code,
            "name": name, "brand": brand,
            "price": price, "qty": qty, "active": active,
        }
    return list(rows.values())

def _parse_excel_smart(raw: bytes) -> List[Dict[str, Any]]:
    if not _OPENPYXL_OK:
        raise HTTPException(500, "openpyxl не установлен на сервере.")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Не удалось открыть Excel: {e}")
    ws = wb.active
    headers = [str(c.value or '').strip() for c in ws[1]]

    def norm(h: str) -> str:
        return "".join(ch for ch in h.lower() if ch.isalnum())

    aliases = {
        "sku": {"sku","code","shopsku","shopski","vendorcode","offerid","id","артикул","код"},
        "name": {"name","model","title","productname","offername","наименование","название","товар"},
        "brand":{"brand","vendor","producer","manufacturer","бренд","производитель"},
        "category":{"category","categoryname","group","группа","категория"},
        "price":{"price","baseprice","saleprice","currentprice","totalprice","cityprice","цена"},
        "qty":{"qty","quantity","stock","stockqty","stockquantity","stockcount","availableamount","остаток","количество","шт"},
        "barcode":{"barcode","ean","штрихкод","баркод"},
        "active":{"active","isactive","ispublished","visible","isvisible","status","опубликован","статус"},
    }
    rev = {k:{norm(x) for x in v} for k,v in aliases.items()}
    col2key: Dict[int,str] = {}
    for i,h in enumerate(headers):
        nh = norm(h)
        for tgt, pool in rev.items():
            if nh in pool:
                col2key[i] = tgt
                break

    out: Dict[str, Dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "", []) for v in row):
            continue
        item: Dict[str, Any] = {}
        for i,val in enumerate(row):
            k = col2key.get(i)
            if not k:
                continue
            if k == "price":
                item[k] = _maybe_float(val)
            elif k == "qty":
                item["qty"] = _maybe_int(val)
            elif k == "active":
                if val is None: item[k] = None
                else:
                    s = str(val).strip().lower()
                    item[k] = True if s in ("1","true","yes","on","да","+","опубликован") else \
                              False if s in ("0","false","no","off","нет","-") else None
            else:
                item[k] = (str(val).strip() if val is not None else None)

        sku = _sku_of(item)
        if not sku:
            continue

        if item.get("active") is None and item.get("qty") is not None:
            try:
                item["active"] = True if int(item["qty"]) > 0 else None
            except Exception:
                pass

        out[sku] = {
            "sku": sku, "code": sku,
            "name": item.get("name"),
            "brand": item.get("brand"),
            "category": item.get("category"),
            "price": item.get("price"),
            "qty": item.get("qty"),
            "active": item.get("active"),
            "barcode": item.get("barcode"),
        }
    return list(out.values())

# ──────────────────────────────────────────────────────────────────────────────
# Import/sync core
# ──────────────────────────────────────────────────────────────────────────────
def _dedupe(items: List[Dict[str,Any]]) -> Tuple[List[Dict[str,Any]], List[str]]:
    uniq: Dict[str,Dict[str,Any]] = {}
    dups: List[str] = []
    for it in items:
        sku = _sku_of(it)
        if not sku:
            continue
        if sku in uniq:
            dups.append(sku)
            continue
        uniq[sku] = it
    return list(uniq.values()), sorted(list(set(dups)))

def _existing_sku_set(skus: List[str]) -> set[str]:
    if not skus:
        return set()
    with _db() as c:
        if _USE_PG:
            placeholders = ", ".join([f":s{i}" for i in range(len(skus))])
            sql = _q(f"SELECT sku FROM products WHERE sku IN ({placeholders})")
            params = {f"s{i}": s for i, s in enumerate(skus)}
            rows = c.execute(sql, params).all()
            return {r._mapping["sku"] for r in rows}
        else:
            placeholders = ", ".join(["?"] * len(skus))
            rows = c.execute(f"SELECT sku FROM products WHERE sku IN ({placeholders})", skus).fetchall()
            return {r["sku"] for r in rows}

def _smart_import_bytes(filename: str, content: bytes, *, city_id: str) -> List[Dict[str,Any]]:
    fn = (filename or "").lower()
    if fn.endswith(".xml"):
        items = _parse_xml_smart(content, city_id=city_id)
    elif fn.endswith(".xlsx") or fn.endswith(".xls"):
        items = _parse_excel_smart(content)
    else:
        raise HTTPException(400, "Поддерживаются XML и Excel (.xlsx/.xls).")
    cleaned, _ = _dedupe(items)
    return cleaned

def _sync_with_file(
    items: List[Dict[str,Any]],
    *, mode: str = "replace",
    only_prices: bool = False,
    hard_delete_missing: bool = False
) -> Dict[str,Any]:
    """
    mode:
      - 'replace' => привести состав БД к файлу (отсутствующие деактивировать/удалить)
      - 'merge'   => ничего не удалять, только upsert
    """
    # upsert
    inserted, updated = _upsert_products(items, price_only=only_prices)

    deactivated = deleted = 0
    if mode.lower() != "merge":
        keep = [_sku_of(x) for x in items if _sku_of(x)]
        if hard_delete_missing:
            deleted = _delete_missing(keep)
        else:
            deactivated = _deactivate_missing(keep)

    return {
        "items_in_file": len(items),
        "inserted": inserted,
        "updated": updated,
        "deactivated": deactivated,
        "deleted": deleted,
    }

# ──────────────────────────────────────────────────────────────────────────────
# Simple inline Kaspi sync (XML feed) — безопасно, без внешних модулей
# ──────────────────────────────────────────────────────────────────────────────
@dataclass
class KaspiSyncResult:
    items_in_kaspi: int
    inserted: int
    updated: int
    in_sale: int
    removed: int
    deactivated: int
    deleted: int
    safety_skipped: bool = False
    reason: Optional[str] = None
    source: str = "none"

def _active_final(item: Dict[str, Any]) -> bool:
    """Итоговая активность по правилам Kaspi/нашей логики:
       True/False → как есть; None → qty>0? иначе дефолт из ENV."""
    if item.get("active") is True:
        return True
    if item.get("active") is False:
        return False
    q = _maybe_int(item.get("qty"))
    if q is not None:
        return q > 0
    return _env_bool("KASPI_DEFAULT_ACTIVE", True)

def _fetch_kaspi_items_via_xml() -> Tuple[List[Dict[str,Any]], str]:
    url = os.getenv("KASPI_PRICE_XML_URL") or ""
    if not url:
        return [], "disabled:no-url"
    if not _REQ_OK:
        raise HTTPException(500, "Для KASPI_PRICE_XML_URL требуется пакет 'requests'. Установите его в образ.")
    try:
        r = requests.get(url, timeout=60)
        r.raise_for_status()
    except Exception as e:
        raise HTTPException(502, f"Не удалось скачать XML-фид Kaspi: {e}")
    city_id = os.getenv("KASPI_CITY_ID", "196220100")
    items = _parse_xml_smart(r.content, city_id=city_id)
    items, _ = _dedupe(items)
    return items, f"xml:{url}"

def _run_kaspi_sync_inline(
    *, mode: str, price_only: bool, hard_delete_missing: bool
) -> KaspiSyncResult:
    # 1) источник
    items: List[Dict[str,Any]] = []
    source = "none"

    # пробуем XML-фид
    try:
        items, source = _fetch_kaspi_items_via_xml()
    except HTTPException as e:
        # отдаём понятную ошибку наверх
        raise
    except Exception as e:
        # неизвестная ошибка скачивания/парсинга
        raise HTTPException(502, f"Ошибка получения фида Kaspi: {e}")

    if not items:
        # источник не настроен — вернуть пустой результат, ничего не трогаем
        return KaspiSyncResult(
            items_in_kaspi=0, inserted=0, updated=0, in_sale=0, removed=0,
            deactivated=0, deleted=0, safety_skipped=False, reason="Kaspi sync source is not configured", source=source
        )

    # 2) safety для replace (минимальная доля позиций)
    safety_skipped = False
    reason = None
    if mode.lower() == "replace":
        min_ratio = _env_float("KASPI_REPLACE_SAFETY_MIN_RATIO", 0.5)
        active_now = _count_active_in_db()
        if active_now > 0 and (len(items) / float(active_now)) < float(min_ratio):
            # слишком мало позиций — пропускаем деактивацию
            safety_skipped = True
            reason = f"skip-deactivate: items_in_kaspi={len(items)} < {min_ratio*100:.0f}% of active_in_db={active_now}"
            # принудительно делаем merge (без удаления)
            mode = "merge"

    # 3) синк в БД
    sync_res = _sync_with_file(
        items, mode=mode, only_prices=price_only, hard_delete_missing=hard_delete_missing
    )

    # 4) подсчёты
    in_sale = sum(1 for it in items if _active_final(it))
    removed = sum(1 for it in items if it.get("active") is False)

    return KaspiSyncResult(
        items_in_kaspi=len(items),
        inserted=int(sync_res["inserted"]),
        updated=int(sync_res["updated"]),
        in_sale=int(in_sale),
        removed=int(removed),
        deactivated=int(sync_res["deactivated"]),
        deleted=int(sync_res["deleted"]),
        safety_skipped=safety_skipped,
        reason=reason,
        source=source,
    )

# ──────────────────────────────────────────────────────────────────────────────
# Pydantic models (batches)
# ──────────────────────────────────────────────────────────────────────────────
class BatchIn(BaseModel):
    date: str
    qty: int
    unit_cost: float
    note: str | None = None
    commission_pct: float | None = None
    batch_code: str | None = None

class BatchListIn(BaseModel):
    entries: List[BatchIn]

# ──────────────────────────────────────────────────────────────────────────────
# FIFO recount helper
# ──────────────────────────────────────────────────────────────────────────────
def _recount_qty_sold_from_ledger() -> int:
    _ensure_schema()
    updated = 0
    with _db() as c:
        ledgers = ["profit_fifo_ledger", "fifo_ledger", "ledger_fifo"]
        ledger = next((t for t in ledgers if _table_exists(c, t)), None)
        if not ledger:
            return 0
        batch_col = next((col for col in ("batch_id","bid","batch") if _has_column(c, ledger, col)), None)
        qty_col   = next((col for col in ("qty","quantity","qty_used") if _has_column(c, ledger, col)), None)
        if not batch_col or not qty_col:
            return 0

        if _USE_PG:
            rows = c.execute(_q(
                f"SELECT {batch_col} AS bid, SUM({qty_col}) AS used FROM {ledger} GROUP BY {batch_col}"
            )).all()
            used = {int(r._mapping["bid"]): int(r._mapping["used"] or 0) for r in rows}
            c.execute(_q("UPDATE batches SET qty_sold = COALESCE(qty_sold,0)"))
            for bid, val in used.items():
                r = c.execute(_q(
                    "UPDATE batches SET qty_sold=:v WHERE id=:bid AND COALESCE(qty_sold,0) <> :v"
                ), {"v": int(val), "bid": int(bid)})
                updated += (r.rowcount or 0)
            c.execute(_q("UPDATE batches SET qty_sold = qty WHERE qty_sold > qty"))
        else:
            rows = c.execute(
                f"SELECT {batch_col} AS bid, SUM({qty_col}) AS used FROM {ledger} GROUP BY {batch_col}"
            ).fetchall()
            used = {int(r["bid"]): int(r["used"] or 0) for r in rows}
            c.execute("UPDATE batches SET qty_sold = COALESCE(qty_sold,0)")
            for bid, val in used.items():
                cur = c.execute(
                    "UPDATE batches SET qty_sold=? WHERE id=? AND COALESCE(qty_sold,0) <> ?",
                    (int(val), int(bid), int(val))
                )
                updated += cur.rowcount or 0
            c.execute("UPDATE batches SET qty_sold = qty WHERE qty_sold > qty")
            _commit(c)
    return updated

# ──────────────────────────────────────────────────────────────────────────────
# Local DB listings helper (fix active filter)
# ──────────────────────────────────────────────────────────────────────────────
def list_from_db(*, active: Optional[bool], limit: int = 1000, offset: int = 0, search: str = "") -> List[Dict[str, Any]]:
    _ensure_schema()
    with _db() as c:
        if _USE_PG:
            sql = "SELECT sku, name, brand, category, price, quantity, active FROM products"
            conds, params = [], {}
            if active is True:
                conds.append("active=1")
            elif active is False:
                conds.append("active=0")
            if search:
                conds.append("(sku ILIKE :q OR name ILIKE :q)")
                params["q"] = f"%{search}%"
            if conds:
                sql += " WHERE " + " AND ".join(conds)
            sql += " ORDER BY name LIMIT :lim OFFSET :off"
            params.update({"lim": limit, "off": offset})
            rows = _rows_to_dicts(c.execute(_q(sql), params).all())
        else:
            sql = "SELECT sku, name, brand, category, price, quantity, active FROM products"
            conds, params = [], []
            if active is True:
                conds.append("active=1")
            elif active is False:
                conds.append("active=0")
            if search:
                conds.append("(sku LIKE ? OR name LIKE ?)")
                params += [f"%{search}%", f"%{search}%"]
            if conds:
                sql += " WHERE " + " AND ".join(conds)
            sql += " ORDER BY name COLLATE NOCASE LIMIT ? OFFSET ?"
            params += [limit, offset]
            rows = [dict(r) for r in c.execute(sql, params).fetchall()]
    return rows

# ──────────────────────────────────────────────────────────────────────────────
# Router
# ──────────────────────────────────────────────────────────────────────────────
def get_products_router(*_, **__) -> APIRouter:
    """
    Возвращает готовый APIRouter для include_router(..., prefix="/products")
    """
    router = APIRouter(tags=["products"])

    # Ping
    @router.get("/db/ping")
    async def db_ping():
        _ensure_schema()
        if _USE_PG:
            with _db() as c:
                c.execute(_q("SELECT 1"))
            return {"ok": True, "driver": "pg"}
        else:
            with _db() as c:
                ok = c.execute("PRAGMA integrity_check").fetchone()[0]
            return {"ok": ok == "ok", "driver": "sqlite"}

    @router.get("/my-skus")
    async def my_skus(kaspi: KaspiClient = Depends(get_kaspi_client)):
    # пример вызова
    # for order in kaspi.iter_orders(...): ...
        return {"ok": True}

    # Список из БД (для таблицы «Мой склад») + пагинация
    @router.get("/db/list")
    async def db_list(
        active_only: int = Query(1),
        search: str = Query("", alias="q"),
        page: int = Query(1, ge=1),
        page_size: int = Query(200, ge=1, le=100_000),
    ):
        _ensure_schema()
        _seed_categories_if_empty()

        # categories
        with _db() as c:
            if _USE_PG:
                cats_rows = c.execute(_q(
                    "SELECT name, base_percent, extra_percent, tax_percent FROM categories ORDER BY name"
                )).all()
                cats = {r._mapping["name"]: dict(r._mapping) for r in cats_rows}
            else:
                rows = c.execute(
                    "SELECT name,base_percent,extra_percent,tax_percent FROM categories ORDER BY name COLLATE NOCASE"
                ).fetchall()
                cats = {r["name"]: dict(r) for r in rows}

        # products
        offset = (page - 1) * page_size
        with _db() as c:
            if _USE_PG:
                sql = "SELECT sku,name,brand,category,price,quantity,active FROM products"
                conds, params = [], {}
                if active_only:
                    conds.append("active=1")
                if search:
                    conds.append("(sku ILIKE :q OR name ILIKE :q)")
                    params["q"] = f"%{search}%"
                if conds:
                    sql += " WHERE " + " AND ".join(conds)
                sql += " ORDER BY name LIMIT :lim OFFSET :off"
                params.update({"lim": page_size, "off": offset})
                rows = _rows_to_dicts(c.execute(_q(sql), params).all())
            else:
                sql = "SELECT sku,name,brand,category,price,quantity,active FROM products"
                conds, params = [], []
                if active_only:
                    conds.append("active=1")
                if search:
                    conds.append("(sku LIKE ? OR name LIKE ?)")
                    params += [f"%{search}%", f"%{search}%"]
                if conds:
                    sql += " WHERE " + " AND ".join(conds)
                sql += " ORDER BY name COLLATE NOCASE LIMIT ? OFFSET ?"
                params += [page_size, offset]
                rows = [dict(r) for r in c.execute(sql, params).fetchall()]

        # meta by batches
        with _db() as c:
            if _USE_PG:
                bc_rows = c.execute(_q("SELECT sku, COUNT(*) AS cnt FROM batches GROUP BY sku")).all()
                bc = {r._mapping["sku"]: r._mapping["cnt"] for r in bc_rows}
                last_rows = c.execute(_q(
                    "SELECT DISTINCT ON (sku) sku, unit_cost, commission_pct FROM batches ORDER BY sku, date DESC, id DESC"
                )).all()
                last = {r._mapping["sku"]: (r._mapping["unit_cost"], r._mapping["commission_pct"]) for r in last_rows}
                left_rows = c.execute(_q(
                    "SELECT sku, SUM(qty - COALESCE(qty_sold,0)) AS left FROM batches GROUP BY sku"
                )).all()
                left_by_sku = {r._mapping["sku"]: int(r._mapping["left"] or 0) for r in left_rows}
            else:
                bc = {r["sku"]: r["cnt"] for r in c.execute(
                    "SELECT sku, COUNT(*) AS cnt FROM batches GROUP BY sku"
                )}
                last = {}
                for r in c.execute(
                    "SELECT sku, unit_cost, commission_pct FROM batches ORDER BY sku, date DESC, id DESC"
                ):
                    if r["sku"] not in last:
                        last[r["sku"]] = (r["unit_cost"], r["commission_pct"])
                left_by_sku = {r["sku"]: int(r["left"] or 0) for r in c.execute(
                    "SELECT sku, SUM(qty - COALESCE(qty_sold,0)) AS left FROM batches GROUP BY sku"
                ).fetchall()}

        items: List[Dict[str, Any]] = []
        for r in rows:
            sku = r["sku"]
            price = float(r.get("price") or 0)
            qty = int(r.get("quantity") or 0)
            cat = r.get("category") or ""
            commissions = cats.get(cat)
            last_margin = None
            if sku in last:
                ucost, comm = last[sku]
                eff_comm = float(comm) if comm is not None else (
                    (float(commissions["base_percent"]) + float(commissions["extra_percent"]) + float(commissions["tax_percent"]))
                    if commissions else 0.0
                )
                last_margin = price - (price * eff_comm/100.0) - float(ucost)

            left_total = int(left_by_sku.get(sku, 0))
            deficit = max(qty - left_total, 0)  # «не хватает партий»

            items.append({
                "code": sku, "id": sku, "name": r.get("name"), "brand": r.get("brand"), "category": cat,
                "qty": qty, "price": price, "active": bool(r.get("active")),
                "batch_count": int(bc.get(sku, 0)),
                "last_margin": round(last_margin, 2) if last_margin is not None else None,
                "left_total": left_total,

                # новое для фронта:
                "deficit": deficit,
                "has_deficit": bool(deficit),
                "in_sale": bool(r.get("active")),  # «в продаже» = активен, независимо от партий
            })
        return {"count": len(items), "items": items}

    # Стоимость остатков (для виджета)
    @router.get("/db/stock-value")
    async def stock_value(with_retail: int = Query(0), details: int = Query(0)):
        _ensure_schema()
        total_cost = 0.0
        total_retail = 0.0
        per_sku: Dict[str, Dict[str, Any]] = {}

        with _db() as c:
            # left * unit_cost
            if _USE_PG:
                rows = c.execute(_q("""
                    SELECT b.sku, SUM(b.qty - COALESCE(b.qty_sold,0)) AS left_qty,
                           SUM( (b.qty - COALESCE(b.qty_sold,0)) * b.unit_cost ) AS cost
                      FROM batches b
                  GROUP BY b.sku
                """)).all()
                left_map = {r._mapping["sku"]: (float(r._mapping["cost"] or 0.0), int(r._mapping["left_qty"] or 0)) for r in rows}
                if with_retail:
                    pr = c.execute(_q("SELECT sku, price FROM products")).all()
                    price_map = {r._mapping["sku"]: float(r._mapping["price"] or 0.0) for r in pr}
                else:
                    price_map = {}
            else:
                rows = c.execute("""
                    SELECT sku, SUM(qty - COALESCE(qty_sold,0)) AS left_qty,
                           SUM( (qty - COALESCE(qty_sold,0)) * unit_cost ) AS cost
                      FROM batches
                  GROUP BY sku
                """).fetchall()
                left_map = {r["sku"]: (float(r["cost"] or 0.0), int(r["left_qty"] or 0)) for r in rows}
                if with_retail:
                    pr = c.execute("SELECT sku, price FROM products").fetchall()
                    price_map = {r["sku"]: float(r["price"] or 0.0) for r in pr}
                else:
                    price_map = {}

            for sku, (cost, left_qty) in left_map.items():
                total_cost += cost
                retail = float(price_map.get(sku, 0.0)) * left_qty if with_retail else 0.0
                total_retail += retail
                if details:
                    per_sku[sku] = {"left_qty": left_qty, "cost": round(cost,2), "retail": round(retail,2)}

        out = {"total_cost": round(total_cost,2)}
        if with_retail:
            out["total_retail"] = round(total_retail,2)
        if details:
            out["items"] = per_sku
        return out

    # Точная карточка товара
    @router.get("/db/sku/{sku}")
    async def get_sku(sku: str):
        _ensure_schema()
        with _db() as c:
            if _USE_PG:
                r = c.execute(_q("SELECT * FROM products WHERE sku=:s"), {"s": sku}).first()
                if not r: raise HTTPException(404, "Not found")
                prod = dict(r._mapping)
                rows = _rows_to_dicts(c.execute(_q(
                    "SELECT id, date, qty, qty_sold, (qty - COALESCE(qty_sold,0)) AS left, unit_cost, commission_pct, batch_code, note "
                    "FROM batches WHERE sku=:s ORDER BY date, id"
                ), {"s": sku}).all())
            else:
                r = c.execute("SELECT * FROM products WHERE sku=?", (sku,)).fetchone()
                if not r: raise HTTPException(404, "Not found")
                prod = dict(r)
                rows = [dict(x) for x in c.execute(
                    "SELECT id, date, qty, qty_sold, (qty - COALESCE(qty_sold,0)) AS left, unit_cost, commission_pct, batch_code, note "
                    "FROM batches WHERE sku=? ORDER BY date, id", (sku,)
                )]
        # avg cost
        with _db() as c:
            if _USE_PG:
                rr = c.execute(_q(
                    "SELECT SUM(qty*unit_cost) AS tc, SUM(qty) AS tq FROM batches WHERE sku=:s"
                ), {"s": sku}).first()
                avgc = None if not rr or not rr._mapping["tq"] else float(rr._mapping["tc"])/float(rr._mapping["tq"])
            else:
                rr = c.execute("SELECT SUM(qty*unit_cost) AS tc, SUM(qty) AS tq FROM batches WHERE sku=?", (sku,)).fetchone()
                avgc = None if not rr or not rr["tq"] else float(rr["tc"])/float(rr["tq"])
        return {"product": prod, "batches": rows, "avg_cost": round(avgc,2) if avgc is not None else None}

    # ── Синхронизация Kaspi (встроенная, через XML фид)
    @router.post("/kaspi/sync", dependencies=[Depends(_require_api_key)])
    async def kaspi_sync_endpoint(
        mode: str = Query("merge", regex="^(merge|replace)$"),
        price_only: bool = Query(True),
        hard_delete_missing: bool = Query(False),
    ):
        res = _run_kaspi_sync_inline(
            mode=mode, price_only=price_only, hard_delete_missing=hard_delete_missing
        )
        if is_dataclass(res):
            return asdict(res)
        return res  # pragma: no cover

    # Совместимость со старым путём
    @router.post("/sync/kaspi/run", dependencies=[Depends(_require_api_key)])
    async def run_kaspi_sync(
        mode: str = Query("merge", regex="^(merge|replace)$"),
        price_only: int = Query(0),
        hard_delete_missing: int = Query(0),
    ):
        res = _run_kaspi_sync_inline(
            mode=mode, price_only=bool(price_only), hard_delete_missing=bool(hard_delete_missing)
        )
        if is_dataclass(res):
            return asdict(res)
        return res

    # Списки (в продаже / снятые)
    @router.get("/db/list/in-sale")
    async def list_in_sale(q: str = Query(""), limit: int = 1000, offset: int = 0):
        return {"items": list_from_db(active=True, limit=limit, offset=offset, search=q)}

    @router.get("/db/list/removed")
    async def list_removed(q: str = Query(""), limit: int = 1000, offset: int = 0):
        return {"items": list_from_db(active=False, limit=limit, offset=offset, search=q)}

    # Массовый upsert (кнопка «Сохранить таблицу в БД»)
    @router.post("/db/bulk-upsert", dependencies=[Depends(_require_api_key)])
    async def bulk_upsert(
        rows: List[Dict[str, Any]] = Body(...),
        price_only: int = Query(1, description="1 — обновлять только цены/мету, qty/active не трогать")
    ):
        if not isinstance(rows, list):
            raise HTTPException(400, "Body должен быть списком объектов")
        inserted, updated = _upsert_products(rows, price_only=bool(price_only))
        return {"ok": True, "inserted": inserted, "updated": updated}

    # Экспорт CSV (из БД)
    @router.get("/db/export.csv")
    async def export_db_csv(active_only: int = Query(1), q: str = Query("", alias="search")):
        _ensure_schema()
        with _db() as c:
            if _USE_PG:
                sql = "SELECT sku,name,brand,category,price,quantity,active,barcode FROM products"
                conds, params = [], {}
                if active_only:
                    conds.append("active=1")
                if q:
                    conds.append("(sku ILIKE :q OR name ILIKE :q)")
                    params["q"] = f"%{q}%"
                if conds:
                    sql += " WHERE " + " AND ".join(conds)
                sql += " ORDER BY name"
                rows = _rows_to_dicts(c.execute(_q(sql), params).all())
            else:
                sql = "SELECT sku,name,brand,category,price,quantity,active,barcode FROM products"
                conds, params = [], []
                if active_only:
                    conds.append("active=1")
                if q:
                    conds.append("(sku LIKE ? OR name LIKE ?)")
                    params += [f"%{q}%", f"%{q}%"]
                if conds:
                    sql += " WHERE " + " AND ".join(conds)
                sql += " ORDER BY name COLLATE NOCASE"
                rows = [dict(r) for r in c.execute(sql, params).fetchall()]

        def esc(s: Any) -> str:
            s = "" if s is None else str(s)
            if any(c in s for c in [",", '"', "\n"]):
                s = '"' + s.replace('"', '""') + '"'
            return s

        header = "sku,name,brand,category,price,quantity,active,barcode\n"
        body = "".join(
            ",".join(esc(x) for x in [
                r["sku"], r["name"], r.get("brand") or "", r.get("category") or "",
                r.get("price") or 0, r.get("quantity") or 0, 1 if r.get("active") else 0, r.get("barcode") or ""
            ]) + "\n"
            for r in rows
        )
        return Response(
            content=header + body,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="products-db.csv"'}
        )

    # ──────────────────────────────────────────────────────────────────────
    # ИМПОРТ / СИНХРОНИЗАЦИЯ
    # ──────────────────────────────────────────────────────────────────────
    @router.post("/import/sync", dependencies=[Depends(_require_api_key)])
    async def import_sync(
        file: UploadFile = File(...),
        mode: str = Query("replace", regex="^(replace|merge)$"),
        only_prices: int = Query(0),
        hard_delete_missing: int = Query(0),
        city_id: str = Query(os.getenv("KASPI_CITY_ID", "196220100")),
        dry_run: int = Query(0),
    ):
        raw = await file.read()
        items = _smart_import_bytes(file.filename or "", raw, city_id=city_id)
        items, duplicates = _dedupe(items)

        if dry_run:
            skus = [_sku_of(x) for x in items]
            exists = _existing_sku_set(skus)
            return {
                "dry_run": True,
                "items_in_file": len(items),
                "inserted": len([s for s in skus if s not in exists]),
                "updated": len([s for s in skus if s in exists]),
                "duplicates": duplicates,
            }

        # safety для replace
        if mode.lower() == "replace":
            min_ratio = _env_float("KASPI_REPLACE_SAFETY_MIN_RATIO", 0.5)
            active_now = _count_active_in_db()
            if active_now > 0 and (len(items) / float(active_now)) < float(min_ratio):
                # пропускаем деактивацию
                res = _sync_with_file(items, mode="merge", only_prices=bool(only_prices), hard_delete_missing=False)
                res.update({
                    "safety_skipped": True,
                    "reason": f"skip-deactivate: items_in_file={len(items)} < {min_ratio*100:.0f}% of active_in_db={active_now}"
                })
                res["duplicates"] = duplicates
                return res

        res = _sync_with_file(
            items,
            mode=mode,
            only_prices=bool(only_prices),
            hard_delete_missing=bool(hard_delete_missing)
        )
        res["duplicates"] = duplicates
        return res

    # Совместимость (старые фронты) — делаем merge-режим по умолчанию
    @router.post("/import", dependencies=[Depends(_require_api_key)])
    async def import_compat(
        file: UploadFile = File(...),
        price_only: int = Query(0),
        city_id: str = Query(os.getenv("KASPI_CITY_ID", "196220100")),
        dry_run: int = Query(0),
    ):
        raw = await file.read()
        items = _smart_import_bytes(file.filename or "", raw, city_id=city_id)
        items, duplicates = _dedupe(items)
        if dry_run:
            skus = [_sku_of(x) for x in items]
            exists = _existing_sku_set(skus)
            return {
                "dry_run": True,
                "items_in_file": len(items),
                "inserted": len([s for s in skus if s not in exists]),
                "updated": len([s for s in skus if s in exists]),
                "duplicates": duplicates,
            }
        res = _sync_with_file(items, mode="merge", only_prices=bool(price_only))
        res["duplicates"] = duplicates
        return res

    @router.post("/manual-upload", dependencies=[Depends(_require_api_key)])
    async def manual_upload(
        file: UploadFile = File(...),
        mode: str = Query("replace"),
        only_prices: int = Query(0),
        city_id: str = Query(os.getenv("KASPI_CITY_ID", "196220100")),
        dry_run: int = Query(0),
    ):
        # пробрасываем на новый синхро-эндпоинт
        raw = await file.read()
        items = _smart_import_bytes(file.filename or "", raw, city_id=city_id)
        items, duplicates = _dedupe(items)
        if dry_run:
            skus = [_sku_of(x) for x in items]
            exists = _existing_sku_set(skus)
            return {
                "dry_run": True,
                "items_in_file": len(items),
                "inserted": len([s for s in skus if s not in exists]),
                "updated": len([s for s in skus if s in exists]),
                "duplicates": duplicates,
            }
        res = _sync_with_file(items, mode=mode, only_prices=bool(only_prices))
        res["duplicates"] = duplicates
        return res

    # ──────────────────────────────────────────────────────────────────────
    # ПАРТИИ
    # ──────────────────────────────────────────────────────────────────────
    @router.get("/db/price-batches/{sku}")
    async def get_batches(sku: str):
        _ensure_schema()
        with _db() as c:
            if _USE_PG:
                rows = _rows_to_dicts(c.execute(_q(
                    "SELECT id, date, qty, qty_sold, (qty - COALESCE(qty_sold,0)) AS left, "
                    "unit_cost, commission_pct, batch_code, note "
                    "FROM batches WHERE sku=:sku ORDER BY date, id"
                ), {"sku": sku}).all())
            else:
                rows = [dict(r) for r in c.execute(
                    "SELECT id, date, qty, qty_sold, (qty - COALESCE(qty_sold,0)) AS left, "
                    "unit_cost, commission_pct, batch_code, note "
                    "FROM batches WHERE sku=? ORDER BY date, id", (sku,)
                )]
        # avg cost
        with _db() as c:
            if _USE_PG:
                r = c.execute(_q(
                    "SELECT SUM(qty*unit_cost) AS tc, SUM(qty) AS tq FROM batches WHERE sku=:s"
                ), {"s": sku}).first()
                avgc = None if not r or not r._mapping["tq"] else float(r._mapping["tc"])/float(r._mapping["tq"])
            else:
                r = c.execute("SELECT SUM(qty*unit_cost) AS tc, SUM(qty) AS tq FROM batches WHERE sku=?", (sku,)).fetchone()
                avgc = None if not r or not r["tq"] else float(r["tc"])/float(r["tq"])
        return {"batches": rows, "avg_cost": round(avgc, 2) if avgc is not None else None}

    @router.post("/db/price-batches/{sku}", dependencies=[Depends(_require_api_key)])
    async def add_batches(sku: str, payload: BatchListIn = Body(...)):
        _ensure_schema()
        # safety: ensure sku exists to avoid FK error (auto-create placeholder)
        with _db() as c:
            if _USE_PG:
                ex = c.execute(_q("SELECT 1 FROM products WHERE sku=:s"), {"s": sku}).first()
            else:
                ex = c.execute("SELECT 1 FROM products WHERE sku=?", (sku,)).fetchone()
        if not ex:
            _upsert_products([{"sku": sku, "name": sku, "price": 0, "qty": 0, "active": 1}], price_only=False)

        with _db() as c:
            for e in payload.entries:
                code = e.batch_code or _gen_batch_code()
                if _USE_PG:
                    c.execute(_q(
                        "INSERT INTO batches(sku,date,qty,unit_cost,note,commission_pct,batch_code,qty_sold) "
                        "VALUES(:sku,:date,:qty,:ucost,:note,:comm,:code,0)"
                    ), {"sku": sku, "date": e.date, "qty": int(e.qty),
                        "ucost": float(e.unit_cost), "note": e.note,
                        "comm": float(e.commission_pct) if e.commission_pct is not None else None,
                        "code": code})
                else:
                    c.execute(
                        "INSERT INTO batches(sku,date,qty,unit_cost,note,commission_pct,batch_code,qty_sold) "
                        "VALUES(?,?,?,?,?,?,?,0)",
                        (sku, e.date, int(e.qty), float(e.unit_cost), e.note,
                         float(e.commission_pct) if e.commission_pct is not None else None, code)
                    )
            _commit(c)
        return {"ok": True}

    @router.put("/db/price-batches/{sku}/{bid}", dependencies=[Depends(_require_api_key)])
    async def update_batch(sku: str, bid: int, payload: Dict[str, Any] = Body(...)):
        fields = {
            "date": payload.get("date"),
            "qty": payload.get("qty"),
            "unit_cost": payload.get("unit_cost"),
            "note": payload.get("note"),
            "commission_pct": payload.get("commission_pct"),
            "batch_code": payload.get("batch_code"),
        }
        sets = {k: v for k, v in fields.items() if v is not None}
        if not sets:
            return {"ok": True, "status": "noop"}
        _ensure_schema()
        with _db() as c:
            if _USE_PG:
                parts = [f"{k}=:{k}" for k in sets.keys()]
                sets["bid"] = bid; sets["sku"] = sku
                c.execute(_q(f"UPDATE batches SET {', '.join(parts)} WHERE id=:bid AND sku=:sku"), sets)
                c.execute(_q("UPDATE batches SET qty_sold = LEAST(qty_sold, qty) WHERE id=:bid AND sku=:sku"),
                          {"bid": bid, "sku": sku})
            else:
                parts = [f"{k}=?" for k in sets.keys()]
                params = list(sets.values()) + [bid, sku]
                c.execute(f"UPDATE batches SET {', '.join(parts)} WHERE id=? AND sku=?", params)
                c.execute("UPDATE batches SET qty_sold = MIN(qty_sold, qty) WHERE id=? AND sku=?", (bid, sku))
                _commit(c)
        return {"ok": True}

    @router.delete("/db/price-batches/{sku}/{bid}", dependencies=[Depends(_require_api_key)])
    async def delete_batch(sku: str, bid: int):
        _ensure_schema()
        with _db() as c:
            if _USE_PG:
                r = c.execute(_q("SELECT COALESCE(qty_sold,0) AS s FROM batches WHERE id=:bid AND sku=:sku"),
                              {"bid": bid, "sku": sku}).first()
                if not r:
                    raise HTTPException(404, "Batch not found")
                if int(r._mapping["s"]) > 0:
                    raise HTTPException(400, "Cannot delete: batch has sales")
                c.execute(_q("DELETE FROM batches WHERE id=:bid AND sku=:sku"), {"bid": bid, "sku": sku})
            else:
                r = c.execute("SELECT COALESCE(qty_sold,0) AS s FROM batches WHERE id=? AND sku=?", (bid, sku)).fetchone()
                if not r:
                    raise HTTPException(404, "Batch not found")
                if int(r["s"]) > 0:
                    raise HTTPException(400, "Cannot delete: batch has sales")
                c.execute("DELETE FROM batches WHERE id=? AND sku=?", (bid, sku))
                _commit(c)
        return {"ok": True}

    @router.post("/batches/recount-sold", dependencies=[Depends(_require_api_key)])
    async def batches_recount_sold():
        changed = _recount_qty_sold_from_ledger()
        return {"ok": True, "updated_batches": int(changed)}

    # История по SKU из profit_fifo_ledger (если есть)
    @router.get("/db/ledger/{sku}")
    async def ledger_by_sku(sku: str, limit: int = Query(200, ge=1, le=2000)):
        _ensure_schema()
        with _db() as c:
            if not _table_exists(c, "profit_fifo_ledger"):
                return {"ok": True, "items": []}
            if _USE_PG:
                rows = _rows_to_dicts(c.execute(_q("""
                    SELECT id, order_code, date_utc_ms, line_index, qty, unit_price, total_price,
                           batch_id, unit_cost, commission_pct, commission_amount, cost_amount, profit_amount
                      FROM profit_fifo_ledger
                     WHERE sku=:sku
                     ORDER BY date_utc_ms DESC, id DESC
                     LIMIT :lim
                """), {"sku": sku, "lim": limit}).all())
            else:
                rows = [dict(r) for r in c.execute("""
                    SELECT id, order_code, date_utc_ms, line_index, qty, unit_price, total_price,
                           batch_id, unit_cost, commission_pct, commission_amount, cost_amount, profit_amount
                      FROM profit_fifo_ledger
                     WHERE sku=?
                     ORDER BY date_utc_ms DESC, id DESC
                     LIMIT ?
                """, (sku, limit)).fetchall()]
        return {"ok": True, "items": rows}

    # Ensure SKU (на лету создать/обновить)
    @router.post("/db/ensure-sku/{sku}", dependencies=[Depends(_require_api_key)])
    async def ensure_sku(
        sku: str,
        name: str = Query(...),
        price: float = Query(0.0),
        qty: int = Query(0),
        active: int = Query(1),
        brand: Optional[str] = None,
        category: Optional[str] = None,
        barcode: Optional[str] = None,
    ):
        payload = [{
            "sku": _norm_sku(sku), "code": _norm_sku(sku), "name": name,
            "price": float(price), "qty": int(qty),
            "active": int(active), "brand": brand,
            "category": category, "barcode": barcode,
        }]
        ins, upd = _upsert_products(payload, price_only=False)
        return {"ok": True, "inserted": ins, "updated": upd}

    # Бэкап/восстановление (SQLite only)
    @router.get("/db/backup.sqlite3")
    async def backup_db():
        _ensure_schema()
        if _USE_PG:
            raise HTTPException(501, "Backup доступен только для локальной SQLite.")
        fname = os.path.basename(DB_PATH) or "data.sqlite3"
        return FileResponse(DB_PATH, media_type="application/octet-stream", filename=fname)

    @router.post("/db/restore")
    async def restore_db(file: UploadFile = File(...)):
        if _USE_PG:
            raise HTTPException(501, "Restore доступен только для локальной SQLite.")
        content = await file.read()
        try:
            with _db() as c:
                c.execute("PRAGMA wal_checkpoint(TRUNCATE);")
        except Exception:
            pass
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
        with open(DB_PATH, "wb") as f:
            f.write(content)
        with _db() as c:
            ok = c.execute("PRAGMA integrity_check").fetchone()[0]
        return {"ok": True, "integrity": ok}

    return router
